import argparse
import logging
from tableaudocumentapi import Datasource
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
# from openpyxl.worksheet.dimensions import ColumnDimension

_INDEX_ID = 0
_INDEX_NAME = 1
_INDEX_CAPTION = 2
_INDEX_TEXT = 3
_INDEX_ROLE = 4
_INDEX_TYPE = 5
_INDEX_PARAM = 7


def main():
    parser = argparse.ArgumentParser(description="""Import/export datasource metadata. Note: delete any extract/hyper files otherwise metadata will be reset again based on the contents of those files.""")
    parser.add_argument('--file', '-f', required=True, help='metadata sheet to use')
    parser.add_argument('-i', required=False, action='store_true', help='import metadata')
    parser.add_argument('--overwrite', '-w', required=False, action='store_true', help='overwrite existing metadata in the datasource(s)')
    parser.add_argument('--use_params', '-u', required=False, action='store_true', help='overwrite existing metadata using param column values')

    parser.add_argument('--logging-level', '-l', choices=['debug', 'info', 'error'], default='error',
                        help='desired logging level (set to error by default)')

    parser.add_argument('datasource', help='one or more datasource to process', nargs='+')

    args = parser.parse_args()

    # Set logging level based on user input, or error by default
    logging_level = getattr(logging, args.logging_level.upper())
    logging.basicConfig(level=logging_level)

    for name in args.datasource:
        ds = Datasource.from_file(name)

        if args.i:
            book = load_workbook(filename=args.file)
            sheet = book["Metadata"]
            for row in sheet.rows:
                if "id" == row[_INDEX_ID].value and "name" == row[_INDEX_NAME].value:     # skip the first row with "id", "name"
                    continue
                try:
                    n = set_values(ds, args, row)
                except KeyError:
                    logging.info("can not find '{0}' in data source".format(n))
                    continue
            ds.process_columns()
            ds.save()
        else:
            book = setup_workbook("Metadata")
            sheet = book.active
            for n in ds.fields:
                text = ''
                field = ds.fields[n]
                if field.description:
                    e = ET.XML(field.description)
                    text = remove_blanks(e)
                if text is None:
                    text = ''
                db_ref = ds.db_columns.get(field.id)
                db_col = "N/A"
                if db_ref:
                    db_col = db_ref.value
                sheet.append((field.id, n, field.caption, text, field.role, field.type, db_col, False))
            book.save(args.file)


def set_values(ds, args, row):
    id = row[_INDEX_ID].value
    n = row[_INDEX_NAME].value
    field = ds.fields[n]
    caption = row[_INDEX_CAPTION].value
    text = row[_INDEX_TEXT].value
    role = row[_INDEX_ROLE].value
    type = row[_INDEX_TYPE].value
    if (text is not None and text != '' and (
            field.description is None or remove_blanks(ET.XML(field.description)) == "")) \
            or (args.overwrite) \
            or (args.use_params and row[_INDEX_PARAM].value):
        if caption is not None:
            field.caption = caption
            logging.debug("id [{2}], field [{0}] overwritten with caption [{1}]".format(n, caption, id))
        if text is not None:
            field.description = str(text)
            logging.debug("id [{2}], field [{0}] overwritten with description [{1}]".format(n, text, id))
        if role is not None:
            field.role = role
        if type is not None:
            field.type = type
    return n


def setup_workbook(name):
    book = Workbook()
    while len(book.worksheets) > 0:
        book.remove_sheet(book.worksheets[0])
    sheet = book.create_sheet(title=name)
    # sheet = book.active
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.fitToWidth = 1
    sheet.append(("id", "name", "caption", "description", "role", "type", "DB", "param"))
    sheet.freeze_panes = 'C2'
    sheet.column_dimensions['A'].hidden = True
    sheet.column_dimensions['B'].hidden = False
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 100
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 100
    sheet.column_dimensions['G'].width = 30
    return book


def remove_blanks(node):
    text = list()
    for x in node:
        if x.tag == 'run' and x.text:
            text.append(x.text)
        else:
            text.append(remove_blanks(x))
    return ''.join(text)


if __name__ == '__main__':
    main()
